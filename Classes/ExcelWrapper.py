from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from typing import Any, Callable
import PIL.Image as PILImage
from pathlib import Path
import csv

class ExcelWrapper:

    def __init__(self, filepath:str|Path):
        '''Takes in a text file and produces and excel file or takes a an excel filepath'''
        if type(filepath) == str: ext = filepath.split(".")[-1]
        elif isinstance(filepath, Path): ext = filepath.suffix
        else: raise TypeError("Incorrect filepath type passed in, only accepts str or Path object.")

            
        if ext == "txt":
            try:
                self.wkbk, self.wkst = self.__create_excel(self.__get_data(filepath, "utf-16", "\t"))
            except UnicodeEncodeError:
                self.wkbk, self.wkst = self.__create_excel(self.__get_data(filepath, "utf-8", "\t"))
        elif ext == "csv":
            self.wkbk, self.wkst = self.__create_excel(self.__get_data(filepath, "utf-8", ","))
        elif ext == "xlsx":
            self.wkbk = load_workbook(filepath)
            self.wkst = self.wkbk.active        
        else:
            raise ValueError(f"Extension: {ext} not supported.")
        
        self.filepath = str(filepath)

    def __get_data(self, filepath:str, encoding:str, delimiter:str )->list[list]:
        '''Gets the data from a utf-16 encoded text file and returns it as a list of lists'''
        data = []
        
        with open(filepath, "r", encoding=encoding) as raw_file:
            raw_data = csv.DictReader(raw_file, delimiter=delimiter)
            for row_dict in raw_data: data.append([val for val in row_dict.values()])
        return [self.__flatten_list(item) for item in data]
    
    def __create_excel(self,data:list[list], col_limit = 14)->tuple[Workbook, Worksheet]:
        '''Appends each list within the list argument to an excel worksheet, passing in a col_limit means that only
        items until the specified index are appended to the excel worksheet'''
        wkbk = Workbook()
        wkst:Worksheet = wkbk.active
        for row in data: wkst.append(row[:col_limit])
        return wkbk, wkst
    
    def __flatten_list(self, lists:list)->list:
        '''flattens inner lists into 1 list, goes 1 level deep only'''
        flattened = []
        for item in lists:
            if type(item) == list:
                for inner_item in item: flattened.append(inner_item)
            else:
                flattened.append(item)
            
        return flattened
    
    def add_column(self, col:str|int, data:str|list, start_row:int = 1)->None:
        '''Adds the data passed in to all the cells in the column, if a single value is passed in all cells
        in that column will have that value starting from start_row'''
        # col = col.upper()
        # col_index = ord(col) - ord("A")+1 if len(col) == 1 else (ord("Z") - ord("A")+1)+ord(col[-1]) - ord("A")+1
        col_index = self.convert_letter_to_col_index(col) if isinstance(col, str) else col
        limit = len(data) if isinstance(data, list) else self.wkst.max_row
        if not isinstance(data, list): data = [data]*limit

        for row, item in zip(range(start_row,limit+start_row), data): 
            self.wkst.cell(row = row, column=col_index, value=item)
        
        return None  
    
    def add_row(self, row:int, data:str|list, start_col:str|int = "A")->int:
        '''Adds the data horizontally beginning at the start column, returns the last column that data was written to'''
        col_index = self.convert_letter_to_col_index(start_col) if isinstance(start_col, str) else start_col
        limit = len(data) if isinstance(data, list) else self.wkst.max_column
        
        for col, item in zip(range(col_index, col_index+limit), data):
            self.wkst.cell(row=row, column=col, value=item)

        return col
    def get_wkst_size(self, dimensions:str)->dict[str, str|int]:
        '''Parses the wkst.dimensions property and returns a dictionary with the information'''
        start, end = dimensions.split(":")

        return {
            "first_col":start[0],
            "first_index":int(start[1:]),
            "last_col":end[0],
            "last_index":int(end[1:]),
        }
    
    def get_matrix(self, top_offset:int, left_offset:int, width:int, height:int, by_col = True, val_only = False)->list[Cell]|list[str]:
        '''Returns a list of cells that is width * height beginning from the starting point, if by_col is set to True
        then the list returned are the cells in the order they appear in the excel file vertically, otherwise list returned are 
        the cells horizontally'''
        cells = []
        for col in range(left_offset, width+left_offset):
            for row in range(top_offset, height+top_offset): 
                cells.append(self.wkst.cell(row, col).value if val_only else self.wkst.cell(row, col))

        return cells
    
    def append_to_cell_value(self, cell:Cell, val:str, postion:str = 'prefix')->None:
        '''Adds another value to the current cell value, doesn't replace the cell value'''
        postion = postion.lower()
        if postion == "prefix": cell.value = f"{val}\t{cell.value}"
        elif postion == "suffix": cell.value = f"{cell.value}\t{val}"
        else: raise ValueError("Only 'prefix' or 'suffix' are acceptable arguments")
        return None

    def replace_cell_value(self, cell:Cell, new_val:str, number_format:str = None)->None:
        '''Replaces the cell value with the new value'''
        if number_format:cell.number_format = number_format
        cell.value = new_val
        return None
    
    def get_cell_value(self, cell:Cell)->Any:
        '''Returns the value within the cell passed in'''
        return cell.value
    
    def get_cell(self, row:int, col:str|int, values_only = False)->Cell|Any:
        col = self.convert_letter_to_col_index(col) if isinstance(col, str) else col
        return self.wkst.cell(row, col).value if values_only else self.wkst.cell(row, col)

    def get_column(self, col:str, start_row = 1, end_row:int = 0, values_only = False, index:int = None)->list[Cell|Any]:
        '''Gets the cells in a specified column in descending order'''
        col_index = self.convert_letter_to_col_index(col) if not index else index
        # col = col.upper()
        # col_index = ord(col) - ord("A")+1 if len(col) == 1 else (ord("Z") - ord("A")+1)+ord(col[-1]) - ord("A")+1
        # print(col_index)
        # return col_index
        
        last_row = end_row if end_row > start_row else self.wkst.max_row
        return [self.wkst.cell(row, col_index) if not values_only else self.wkst.cell(row, col_index).value for row in range(start_row,last_row)]

    def apply_to_adj_columns(self, cell_matrix:list[Cell], foo:Callable, col_height = 8)->list[str|int|float]:
        '''Applies the function to each cell value and its adjacent cell value in the cell matrix, its adjacent is calculated
        by adding the col_height to the first index'''
        results = []
        finished_indices = []
        for index in range(len(cell_matrix)):
            if index in finished_indices: continue
            first_col = index
            second_col = index+col_height
            finished_indices.append(first_col)
            finished_indices.append(second_col)
            first_val = float(cell_matrix[first_col].value)
            second_val = float(cell_matrix[second_col].value)
            results.append(foo(first_val, second_val))
        return results
    
    def get_headers(self, values_only:bool = False)->list[Cell]|list[Any]:
        '''Returns the headers of the current excel worksheet either as mutable cells or just the values within the cells'''
        first_row = self.wkst.iter_rows(min_row=1, max_row=1).__next__()
        return [cell.value for cell in first_row] if values_only else [cell for cell in first_row]
        
    def get_column_by_header(self, header:str, start_row:int = 1, end_row:int = 0, values_only = False)->list[Cell]|list[Any]:
        '''Parses the headers of the excel sheet to find the header with the specified name and returns the cells or values contained in the cells in the column from top to bottom'''
        headers = self.get_headers(values_only=True)
        index = headers.index(header)+1 #Adds one because of openpyxl 1-index system
        return self.get_column(None,start_row=start_row, end_row=end_row, values_only=values_only, index=index) if values_only else self.get_column(None,start_row=start_row, end_row=end_row, values_only=values_only, index=index) 

    def convert_col_index_to_letter(self, index:int)->str:
        '''Converts a column index used by openpyxl to a column letter that can be used to access the data in the specified column'''
        return (index-1) + ord("A")
    
    def convert_letter_to_col_index(self, col:str)->int:
        '''Converts a column letter(s) such as "A" or "AA" into a numerical index value used by openpyxl to access the column in the worksheet'''
        col = col.upper()
        #Since openpxyl uses a 1-index system, the value of 26 is obtained from the ascii value of Z subtracted by the ascii value of A and adding 1, this obtains the index for the "Z" column
        index = 0
        #Use base 26 number system -> First determine to what power to raise 26 to based on the length of the string, aka column, "A", "AA", "AAA", etc the greater the index the greater the power
        base = 26 
        powers = [n for n in range(len(col))]
        powers.reverse()
        #Add the values to the current value of index 
        for char, power in zip(col, powers): 
            index+=(base**power)*(ord(char)-ord("A")+1)
        return index

    def save_as_excel(self, destination:str = None)->None:
        if destination == None:
            destination = Path(self.filepath).with_suffix(".xlsx")
        self.wkbk.save(destination)
        return None
    
    def save_as_csv(self, destination:str,  wkst_name:str = None)->None:
        if wkst_name is None:
            wkst = self.wkbk.active
        else:
            wkst = self.wkbk[wkst_name]
        with open(destination, mode="w", newline="") as csvfile:        
            writer = csv.writer(csvfile, delimiter=",")            
            for row in wkst.iter_rows():
                writer.writerow([cell.value for cell in row if len(row) > 0])
    
    def save_image(self,image:PILImage.Image, cell_coord:str)->None:
        self.wkst.add_image(Image(image), cell_coord)
        return None
    
    def get_last_row(self)->int:
        return self.wkst.max_row
    
    def get_last_col(self)->int:
        return self.wkst.max_column
    
    @classmethod
    def new_excel(cls, wkbk_name:str, wkst_name:str):
        wkbk = Workbook()
        # wkst = wkbk.create_sheet(wkst_name)
        wkbk.save(wkbk_name)
        return cls(wkbk_name)
    
    def __str__(self)->str:
        wkst = self.wkbk.active
        if wkst is None:
            wkst = self.wkbk[self.wkbk.sheetnames[0]]
        for row in wkst.iter_rows(max_row=5):
            print([cell.value for cell in row if len(row) > 0])
        return ""

    def get_rows(self, num:int = None, values_only:bool = True)->list[Any|Cell]:
        wkst = self.wkbk.active
        rows = []
        if wkst is None:
            wkst = self.wkbk[self.wkbk.sheetnames[0]]
        for row in wkst.iter_rows(max_row=num):
            rows.append([cell.value if cell.value is not None else "" for cell in row if len(row) > 0] if values_only else [cell for cell in row if len(row) > 0])
        return rows